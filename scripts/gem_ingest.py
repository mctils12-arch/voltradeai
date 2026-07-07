#!/usr/bin/env python3
"""
gem_ingest.py — Global Energy Monitor asset-registry ingest (DATACORE
MAXIMUS census #5, unblocked 2026-07-07: Mike completed GEM's download
form and delivered the release files; wishlist 9b resolved).

License (verified from the files' own Copyright sheets): CC BY 4.0 —
"Copyright © Global Energy Monitor … Distributed under a Creative
Commons Attribution 4.0 International License." Commercial reuse OK
with attribution. Attribution: "Global Energy Monitor" + per-artifact
recommended citation carried in each output's provenance block.

WHY (hypotheses, gate-locked):
 - Coal mine boundaries + methane sources: geometry for satellite joins
   (FIRMS thermal x mine boundary = activity proxy; S1/S2 chips over
   pits) and the mine -> owner -> parent join into the entity graph.
 - Gas pipelines registry (4,246 rows, NO route geometry in this
   variant — the registry fields only): capacity/status/FID by
   owner/parent — capex-cycle and chokepoint features; parent-company
   names are the equity join spine.
 - Gas finance tracker (LNG terminals + gas power plants with
   financiers and FID status): project-finance cycle features.

Inputs are the GEM release files (paths passed as args — the upload
location is session-specific). Outputs are git-versioned artifacts
under datacore/gem/ (eiaweekly/jodi precedent: whole-file rebuild per
release; git history keeps vintages):
 - coal_mines.json           per-mine registry (centroid/bbox/owners/…)
 - coal_mine_features.geojson.gz   full master FeatureCollection
 - gas_pipelines.json        full 53-field registry, nulls dropped
 - gas_finance.json          LNG terminals + gas power plants sheets

Run (session-side, on each new GEM release):
  python3 scripts/gem_ingest.py <coal_zip> <pipelines_xlsx> <finance_xlsx>
Deps: openpyxl (xlrd/openpyxl precedent from eia_weekly.py).
"""
import gzip
import io
import json
import os
import sys
import zipfile
from datetime import datetime, timezone

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "datacore", "gem")
ATTRIBUTION = "Global Energy Monitor"
LICENSE = "CC BY 4.0 (verified from release Copyright sheets)"


# ── coal mines (master geojson inside the release zip) ──────────────────────

def find_master_geojson(z: zipfile.ZipFile) -> str:
    masters = [n for n in z.namelist()
               if n.startswith("Coal Mine Boundaries") and n.endswith(".geojson")]
    if len(masters) != 1:
        raise RuntimeError(f"expected exactly one master geojson, found {masters!r}")
    return masters[0]


def _walk_coords(geom, acc):
    t = geom.get("type")
    c = geom.get("coordinates")
    if t == "Point":
        acc.append(c)
    elif t in ("MultiPoint", "LineString"):
        acc.extend(c)
    elif t in ("MultiLineString", "Polygon"):
        for part in c:
            acc.extend(part)
    elif t == "MultiPolygon":
        for poly in c:
            for ring in poly:
                acc.extend(ring)
    elif t == "GeometryCollection":
        for g in geom.get("geometries", []):
            _walk_coords(g, acc)


def summarize_coal_mines(master: dict) -> dict:
    """Master FeatureCollection -> per-mine registry. Centroid is the
    bbox midpoint of ALL the mine's feature coordinates (a display
    anchor, NOT a surveyed point — labeled as such)."""
    mines: dict = {}
    for f in master.get("features", []):
        p = f.get("properties", {})
        mid = p.get("GEM Mine ID")
        if not mid:
            continue
        m = mines.setdefault(mid, {
            "name": p.get("Mine Name") or (p.get("Full mine name") or "").split(",")[0],
            "country": p.get("Country / Area"),
            "coal_grade": p.get("Coal Grade"),
            "owners": p.get("Owners"),
            "parent": p.get("Parent Company"),
            "wiki": p.get("GEM Wiki Page (ENG)"),
            "n_features": 0,
            "categories": {},
            "_coords": [],
        })
        m["n_features"] += 1
        cat = p.get("mine feature category") or "unknown"
        sub = p.get("mine feature subcategory")
        key = f"{cat}/{sub}" if sub else cat
        m["categories"][key] = m["categories"].get(key, 0) + 1
        if f.get("geometry"):
            _walk_coords(f["geometry"], m["_coords"])
    for mid, m in mines.items():
        cs = m.pop("_coords")
        if cs:
            lons = [c[0] for c in cs]
            lats = [c[1] for c in cs]
            m["bbox"] = [min(lons), min(lats), max(lons), max(lats)]
            m["centroid_note"] = "bbox midpoint of all features — display anchor, not a surveyed point"
            m["centroid"] = [round((min(lons) + max(lons)) / 2, 5), round((min(lats) + max(lats)) / 2, 5)]
        else:
            m["bbox"] = None
            m["centroid"] = None
    return mines


# ── xlsx sheets -> row dicts (nulls dropped per row) ────────────────────────

def sheet_rows(ws, header_row=1) -> list:
    rows = ws.iter_rows(values_only=True)
    header = None
    out = []
    for i, r in enumerate(rows, start=1):
        if i < header_row:
            continue
        if header is None:
            header = [str(c) if c is not None else f"col{j}" for j, c in enumerate(r)]
            continue
        if all(c is None for c in r):
            continue
        d = {}
        for k, v in zip(header, r):
            if v is None:
                continue
            if isinstance(v, datetime):
                v = v.date().isoformat()
            d[k] = v
        if d:
            out.append(d)
    return out


def provenance(release: str, extra: dict | None = None) -> dict:
    return {
        "attribution": ATTRIBUTION,
        "license": LICENSE,
        "release": release,
        "built_at": datetime.now(timezone.utc).isoformat(),
        **(extra or {}),
    }


def main() -> int:
    if len(sys.argv) != 4:
        print("usage: gem_ingest.py <coal_zip> <pipelines_xlsx> <finance_xlsx>")
        return 2
    coal_zip, pipe_xlsx, fin_xlsx = sys.argv[1:4]
    import openpyxl  # session-side dep

    os.makedirs(OUT_DIR, exist_ok=True)

    # coal mines
    z = zipfile.ZipFile(coal_zip)
    master_name = find_master_geojson(z)
    master = json.loads(z.read(master_name))
    mines = summarize_coal_mines(master)
    if not mines:
        raise RuntimeError("zero mines parsed — refusing to write")
    with gzip.open(os.path.join(OUT_DIR, "coal_mine_features.geojson.gz"), "wt") as f:
        json.dump(master, f, separators=(",", ":"))
    with open(os.path.join(OUT_DIR, "coal_mines.json"), "w") as f:
        json.dump({
            "provenance": provenance(master_name.replace(".geojson", ""),
                                     {"citation": "Global Energy Monitor, Coal Mine Boundaries and Methane Sources"}),
            "mine_count": len(mines),
            "feature_count": len(master.get("features", [])),
            "mines": dict(sorted(mines.items())),
        }, f, separators=(",", ":"))
    print(f"coal: {len(mines)} mines, {len(master.get('features', []))} features")

    # gas pipelines registry
    wb = openpyxl.load_workbook(pipe_xlsx, read_only=True)
    pipes = sheet_rows(wb["Pipelines"])
    if not pipes:
        raise RuntimeError("zero pipeline rows — refusing to write")
    with open(os.path.join(OUT_DIR, "gas_pipelines.json"), "w") as f:
        json.dump({
            "provenance": provenance("Global Gas Infrastructure Tracker - Gas Pipelines - November 2025 release",
                                     {"citation": "Global Energy Monitor, Global Gas Infrastructure Tracker",
                                      "geometry_note": "this release variant carries NO route coordinates — registry fields only (RouteAccuracy/RouteType describe the separate geometry product)"}),
            "row_count": len(pipes),
            "pipelines": pipes,
        }, f, separators=(",", ":"))
    wb.close()
    print(f"pipelines: {len(pipes)} rows")

    # gas finance tracker
    wb = openpyxl.load_workbook(fin_xlsx, read_only=True)
    lng = sheet_rows(wb["LNG Terminals"])
    plants = sheet_rows(wb["Gas Power Plants"])
    if not lng and not plants:
        raise RuntimeError("zero finance rows — refusing to write")
    with open(os.path.join(OUT_DIR, "gas_finance.json"), "w") as f:
        json.dump({
            "provenance": provenance("Gas Finance Tracker - December 2025",
                                     {"citation": "Global Energy Monitor, Gas Finance Tracker"}),
            "lng_terminal_count": len(lng),
            "gas_plant_count": len(plants),
            "lng_terminals": lng,
            "gas_power_plants": plants,
        }, f, separators=(",", ":"))
    wb.close()
    print(f"finance: {len(lng)} LNG terminal rows, {len(plants)} gas plant rows")

    for fn in ("coal_mines.json", "coal_mine_features.geojson.gz", "gas_pipelines.json", "gas_finance.json"):
        fp = os.path.join(OUT_DIR, fn)
        print(f"  {os.path.getsize(fp)/1e6:6.2f}MB  {fn}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
