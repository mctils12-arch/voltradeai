#!/usr/bin/env python3
"""
gem_suite_ingest.py — GEM tracker SUITE ingest (census #5 part 2;
gem-data.zip delivered via Mike's Drive 2026-07-07, 34 members, full
tracker catalog). Every sheet contract below comes from the structure
probe (subagent report 2026-07-07) — header-row exceptions, truncated
sheet names, combined-Coordinates columns and all.

License: CC BY 4.0 per release (verified pattern across the suite);
attribution "Global Energy Monitor".

INCLUSION DECISIONS (stated, never silent — REASONING: git artifacts
must stay clonable; Mike's Drive holds the raw zip as the durable
full-fidelity copy):
 - GEOT (ownership hub): entities + entity->entity edges +
   asset->owner edges INGESTED (nulls dropped, gz over 4MB). The nine
   per-tracker ROLLUP sheets (~240k rows) are SKIPPED: they are the
   transitive closure of nodes+edges — recomputable views, not
   ingredients (archive-ingredients doctrine).
 - GIPT (182,428 power units, exact union of the 8 standalone power
   trackers — CONFIRMED by per-type row counts): ingested as a STATED
   12-field PROJECTION (type, ids, name, country, capacity, status,
   start year, lat/lon, owner/parent entity ids) — full 52-col detail
   stays in the raw zip. The 8 standalone power trackers are SKIPPED
   as exact duplicates; KNOWN DROP: GCPT-only CO2/heat-rate columns
   (in the raw zip when needed).
 - Portal Energetico: SKIPPED — Latin-America regional extract of the
   other trackers (pure duplication, verified).
 - Coal-mine supplements (Sep/Dec 2024): SKIPPED — superseded by GCMT
   May-2026's own Historic Production sheet.
 - GIS zips (68-72MB pipeline route geojsons): NOT ingested here —
   filed as the pipelines map-layer build (PMTiles pipeline).
 - Everything else: full rows, nulls dropped per row, gz when the
   artifact exceeds 4MB plain.

Run: python3 scripts/gem_suite_ingest.py <gem-data.zip>
Deps: openpyxl.
"""
import gzip
import io
import json
import os
import sys
import zipfile
from datetime import datetime, date, timezone

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "datacore", "gem")
GZ_THRESHOLD = 4 * 1024 * 1024
ATTRIBUTION = "Global Energy Monitor"
LICENSE = "CC BY 4.0 (per-release Copyright sheets)"

# (zip member, artifact name, [(sheet, header_row, out_key)...])
# Sheet names EXACT per the probe — including Excel-truncated ones and
# the GOGET trailing-space sheet.
FILE_SPECS = [
    ("Global Coal Mine Tracker, May 2026__.xlsx", "coal_mine_tracker", [
        ("Non-closed mines", 1, "non_closed"),
        ("Closed mines", 1, "closed"),
        ("Historic Production (2018-2025)", 1, "historic_production"),
    ]),
    ("Global-Coal-Terminals-Tracker-December-2024.xlsx", "coal_terminals", [
        ("Terminals", 1, "terminals"),
    ]),
    ("Global-Coal-Project-Finance-Tracker_November-2025_Global-Energy-Monitor-1.xlsx", "coal_finance", [
        ("Data", 2, "financings"),  # header on ROW 2 (row 1 is a banner)
    ]),
    ("Global-Oil-and-Gas-Extraction-Tracker-March-2026.xlsx", "oil_gas_extraction", [
        ("Field-level main data", 1, "fields"),
        ("Field-level reserves data", 1, "field_reserves"),
        ("Field-level production data", 1, "field_production"),
        ("Project-level main data", 1, "projects"),
        ("Project-level reserves data ", 1, "project_reserves"),  # trailing space is REAL
        ("Project-level production data", 1, "project_production"),
    ]),
    ("GEM-GOIT-Oil-NGL-Pipelines-2026-06.xlsx", "oil_ngl_pipelines", [
        ("Data", 1, "pipelines"),
    ]),
    ("GEM-GGIT-LNG-Teminals-2025-09.xlsx", "lng_terminals", [  # GEM's own filename typo
        ("LNG Terminals", 1, "terminals"),
    ]),
    ("LNG-Carrier-Tracker-December-2025-release.xlsx", "lng_carriers", [
        ("data", 1, "carriers"),  # IMO number — joins our AIS archive
    ]),
    ("GMET_V3_12-12-2025.xlsx", "methane_emitters", [
        ("Pipelines", 1, "pipelines"),
        ("LNG Terminals", 1, "lng_terminals"),
        ("Coal Mines - Non-closed", 1, "coal_mines"),
        ("Plumes", 1, "plumes"),  # satellite methane plume OBSERVATIONS (dated events)
        ("Oil and Gas Extraction Areas", 1, "extraction_areas"),
        ("Oil and Gas Reserves", 1, "reserves"),
    ]),
    ("Plant-level_data_Global_Iron_and_Steel_Tracker_June_2026_V1.xlsx", "iron_steel_plants", [
        ("Plant data", 1, "plants"),
        ("Plant capacities and status", 1, "capacities"),
        ("Plant production", 1, "production"),  # float year headers (2019.0)
    ]),
    ("Iron_unit_data_Global_Iron_and_Steel_Tracker_June_2026_V1.xlsx", "iron_units", [
        ("Blast furnaces", 1, "blast_furnaces"),
        ("Direct reduced iron furnaces", 1, "dri_furnaces"),
        ("Blast furnace relinings", 1, "relinings"),
    ]),
    ("Steel_unit_data_Global_Iron_and_Steel_Tracker_June_2026_V1.xlsx", "steel_units", [
        ("Electric arc furnaces", 1, "eaf"),
        ("Basic oxygen furnaces", 1, "bof"),
        ("Induction furnaces", 1, "induction"),
        ("Open hearth furnaces", 1, "open_hearth"),
    ]),
    ("Global-Cement-and-Concrete-Tracker_July-2025.xlsx", "cement", [
        ("Plant Data", 1, "plants"),  # combined "Coordinates" column
    ]),
    ("Global-Iron-Ore-Mines-Tracker-August-2025-V1.xlsx", "iron_ore_mines", [
        ("Main Data", 1, "mines"),  # combined Coordinates
    ]),
    ("Plant-level-data-Global-Chemicals-Inventory-November-2025-V1.xlsx", "chemicals", [
        ("Plant data", 1, "plants"),  # combined Coordinates
    ]),
    ("Production-Consumption-of-Met-Coal-Iron-Ore-by-Steel-Industry-December-2025-Standard-Copy-V1.xlsx", "steel_raw_materials", [
        ("ProdConsMetCoalIronOre", 1, "country_balance"),
    ]),
    ("Global-Energy-Ownership-Tracker-May-2026-V1.xlsx", "ownership", [
        ("All Entities", 1, "entities"),          # LEI/PermID/SEC-CIK crosswalk
        ("Entity Ownership", 1, "entity_edges"),  # entity->entity edges
        ("Asset Ownership", 1, "asset_edges"),    # asset->immediate-owner edges
        # nine rollup sheets SKIPPED — recomputable transitive closure
    ]),
]

# GIPT projection (stated 12-field selection; full 52 cols in raw zip)
GIPT_MEMBER = "Global-Integrated-Power-March-2026-II.xlsx"
GIPT_SHEET = "Power facilities"
GIPT_FIELDS = [
    "Type", "GEM unit/phase ID", "GEM location ID", "Plant / Project name",
    "Country/area", "Capacity (MW)", "Status", "Start year",
    "Latitude", "Longitude", "Owner(s) GEM Entity ID", "Parent(s) GEM Entity ID",
]

SKIPPED = {
    "Portal_Energetico_tracker_2026-06-24.xlsx": "Latin-America regional extract — pure duplication of other trackers (verified)",
    "Global-Coal-Mine-Tracker-September-2024-Supplement-v2.xlsx": "superseded by GCMT May-2026 Historic Production sheet",
    "Global-Coal-Mine-Tracker-December-2024-Supplement-Historical-Production-from-2018-to-2023.xlsx": "superseded by GCMT May-2026 Historic Production sheet",
    "Global-Coal-Plant-Tracker-January-2026.xlsx": "exact subset of GIPT projection (KNOWN DROP: GCPT-only CO2/heat-rate cols — in raw zip)",
    "Global-Solar-Power-Tracker-February-2026.xlsx": "in GIPT projection (distributed-solar country aggregates dropped, stated)",
    "Global-Wind-Power-Tracker-February-2026.xlsx": "in GIPT projection",
    "Global-Hydropower-Tracker-March-2026.xlsx": "in GIPT projection",
    "Global-Oil-and-Gas-Plant-Tracker-GOGPT-January-2026.xlsx": "in GIPT projection",
    "Global-Bioenergy-Power-Tracker-GBPT-V3.xlsx": "in GIPT projection",
    "Global-Nuclear-Power-Tracker-September-2025.xlsx": "in GIPT projection",
    "Geothermal-Power-Tracker-March-2026-Final.xlsx": "in GIPT projection",
}


def norm_cell(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


def sheet_rows(ws, header_row=1):
    """Rows as dicts; header at header_row (verified per sheet); nulls
    dropped; duplicate/blank headers positional-suffixed."""
    it = ws.iter_rows(values_only=True)
    header = None
    out = []
    for i, r in enumerate(it, start=1):
        if i < header_row:
            continue
        if header is None:
            seen: dict = {}
            header = []
            for j, c in enumerate(r):
                name = str(norm_cell(c)) if c is not None else f"col{j}"
                if name in seen:
                    seen[name] += 1
                    name = f"{name}.{seen[name]}"
                else:
                    seen[name] = 0
                header.append(name)
            continue
        if all(c is None for c in r):
            continue
        d = {k: norm_cell(v) for k, v in zip(header, r) if v is not None}
        if d:
            out.append(d)
    return out


def write_artifact(name: str, doc: dict) -> str:
    payload = json.dumps(doc, separators=(",", ":"), default=str)
    plain = os.path.join(OUT_DIR, f"{name}.json")
    if len(payload) > GZ_THRESHOLD:
        fp = plain + ".gz"
        with gzip.open(fp, "wt") as f:
            f.write(payload)
        # a stale plain twin must never shadow the gz
        if os.path.exists(plain):
            os.unlink(plain)
    else:
        fp = plain
        with open(fp, "w") as f:
            f.write(payload)
        if os.path.exists(plain + ".gz"):
            os.unlink(plain + ".gz")
    return fp


def provenance(release: str) -> dict:
    return {
        "attribution": ATTRIBUTION,
        "license": LICENSE,
        "release": release,
        "built_at": datetime.now(timezone.utc).isoformat(),
        "raw_source": "gem-data.zip (Mike's Drive — durable full-fidelity copy)",
    }


def main() -> int:
    if len(sys.argv) != 2:
        print("usage: gem_suite_ingest.py <gem-data.zip>")
        return 2
    import openpyxl
    z = zipfile.ZipFile(sys.argv[1])
    os.makedirs(OUT_DIR, exist_ok=True)
    total = 0

    for member, artifact, sheets in FILE_SPECS:
        wb = openpyxl.load_workbook(io.BytesIO(z.read(member)), read_only=True)
        doc: dict = {"provenance": provenance(member)}
        counts = {}
        for sheet, hdr, key in sheets:
            rows = sheet_rows(wb[sheet], hdr)
            if not rows:
                raise RuntimeError(f"{member} [{sheet}]: zero rows — refusing to write")
            doc[key] = rows
            counts[key] = len(rows)
        doc["counts"] = counts
        wb.close()
        fp = write_artifact(artifact, doc)
        total += os.path.getsize(fp)
        print(f"{artifact:24} {os.path.getsize(fp)/1e6:6.2f}MB  {counts}")

    # GIPT projection
    wb = openpyxl.load_workbook(io.BytesIO(z.read(GIPT_MEMBER)), read_only=True)
    ws = wb[GIPT_SHEET]
    it = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(it)]
    idx = []
    for f in GIPT_FIELDS:
        if f not in header:
            raise RuntimeError(f"GIPT projection field missing from sheet: {f!r}")
        idx.append(header.index(f))
    units = []
    for r in it:
        if all(c is None for c in r):
            continue
        units.append([norm_cell(r[i]) for i in idx])
    if len(units) < 150_000:
        raise RuntimeError(f"GIPT parsed only {len(units)} units — expected ~182k")
    wb.close()
    doc = {
        "provenance": provenance(GIPT_MEMBER),
        "projection_note": "STATED 12-field projection of the 52-col sheet (full detail in raw zip); GIPT is the verified exact union of the 8 standalone power trackers, which are therefore skipped",
        "fields": GIPT_FIELDS,
        "unit_count": len(units),
        "units": units,
    }
    fp = write_artifact("power_units", doc)
    total += os.path.getsize(fp)
    print(f"{'power_units':24} {os.path.getsize(fp)/1e6:6.2f}MB  units={len(units)}")

    with open(os.path.join(OUT_DIR, "suite_skips.json"), "w") as f:
        json.dump({"provenance": provenance("suite-level decisions"), "skipped": SKIPPED,
                   "geot_rollups": "nine per-tracker ownership rollup sheets skipped — recomputable transitive closure of entities+edges"}, f, indent=1)
    print(f"TOTAL artifacts: {total/1e6:.1f}MB")
    return 0


if __name__ == "__main__":
    sys.exit(main())
