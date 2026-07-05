#!/usr/bin/env python3
"""
stb_rail.py — STB EP724 weekly rail service data (BUILD ORDER 2 #6,
2026-07-05; format verified at build per the build-order flag). US
government work, public domain; attribution "Surface Transportation
Board (EP 724 rail service data)".

The consolidated workbook republishes the FULL weekly history since
2017-03 every week (~484 week columns), so this is a whole-file rebuild
each run — git history keeps vintages, and a weekly run appends exactly
one column of new data.

SELECTION (stated here and in the manifest — never silent): of the ~20
EP724 measures, v1 keeps the three that serve the hypothesis (carload
deltas by commodity lead rail earnings + the industrial regime) plus its
service-stress companions:
  - cat 11 "Weekly Carloads By 22 Commodity Categories" (all railroads,
    all 22 commodity variables) — the volume spine
  - cat 1  "Average Train Speed (MPH)", System variable only
  - cat 3  "Cars On Line (Count)", all car-type variables
Everything else (terminal dwell, grain metrics, Chicago gateway, ...) is
deliberately not archived in v1; the workbook is public and re-pullable.

Output: datacore/rail/ep724_carloads.json — one shared week axis + keyed
series ("RAILROAD|MEASURE|VARIABLE" -> value array aligned to the axis,
null where the railroad did not report).

Run (session-side; dep openpyxl — never a runtime dep):
  python3 scripts/stb_rail.py
"""
import json
import os
import re
import urllib.request
from datetime import datetime, timezone

PAGE = "https://www.stb.gov/reports-data/rail-service-data/"
UA = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
OUT = os.path.join(os.path.dirname(__file__), "..", "datacore", "rail", "ep724_carloads.json")

# (category-number-as-string, measure-prefix, variable filter or None=all)
SELECT = [
    ("11", "Weekly Carloads By 22 Commodity Categories", None),
    ("1", "Average Train Speed", {"System"}),
    ("3", "Cars On Line", None),
]


def find_consolidated_url(html):
    """The 'EP724 Consolidated Data through YYYY-MM-DD.xlsx' link; newest
    date wins if several are present."""
    urls = re.findall(r'href="(https://[^"]*EP724%20Consolidated%20Data%20through%20([0-9-]+)\.xlsx)"', html)
    if not urls:
        raise SystemExit("EP724 consolidated link not found — page format changed, refuse to guess")
    return max(urls, key=lambda u: u[1])[0]


def selected(cat, measure, variable):
    m = (measure or "").strip()
    v = (variable or "").strip()
    for c, prefix, vars_ in SELECT:
        if str(cat).strip() == c and m.startswith(prefix):
            return vars_ is None or v in vars_
    return False


def parse_workbook(sheet_rows):
    """Iterator of row tuples (header first) -> {weeks, series}. Values
    coerce to float; non-numeric cells (railroads report '*'/blank for
    not-applicable) become null, never zero."""
    it = iter(sheet_rows)
    header = next(it)
    weeks = []
    for c in header[6:]:
        if isinstance(c, datetime):
            weeks.append(c.date().isoformat())
        elif isinstance(c, str) and re.match(r"\d{4}-\d{2}-\d{2}", c):
            weeks.append(c[:10])
        else:
            weeks.append(None)
    # trailing junk columns (None) are trimmed; interior Nones refuse
    while weeks and weeks[-1] is None:
        weeks.pop()
    if any(w is None for w in weeks):
        raise SystemExit("non-date interior week column — format changed, refuse to guess")
    series = {}
    for row in it:
        if not row or not row[0]:
            continue
        rr, cat, measure, variable = row[0], row[1], row[3], row[4]
        if not selected(cat, measure, variable):
            continue
        key = f"{str(rr).strip()}|{(measure or '').strip()}|{(variable or '').strip()}"
        vals = []
        for c in row[6:6 + len(weeks)]:
            if isinstance(c, (int, float)):
                vals.append(round(float(c), 2))
            else:
                try:
                    vals.append(round(float(str(c).replace(",", "")), 2))
                except (TypeError, ValueError):
                    vals.append(None)
        vals += [None] * (len(weeks) - len(vals))
        series[key] = vals
    return {"weeks": weeks, "series": series}


def main():
    import openpyxl
    req = urllib.request.Request(PAGE, headers={"User-Agent": UA})
    html = urllib.request.urlopen(req, timeout=60).read().decode("utf-8", "replace")
    url = find_consolidated_url(html)
    print("workbook:", url.split("/")[-1])
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    tmp = OUT + ".tmp.xlsx"  # openpyxl validates by extension — must end .xlsx
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(tmp, "wb") as fh:
        fh.write(urllib.request.urlopen(req, timeout=180).read())
    wb = openpyxl.load_workbook(tmp, read_only=True)
    parsed = parse_workbook(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    os.unlink(tmp)
    doc = {
        "built": datetime.now(timezone.utc).isoformat(),
        "source": "STB EP724 consolidated rail service data (public domain)",
        "attribution": "Surface Transportation Board (EP 724 rail service data)",
        "selection": "cat 11 weekly carloads (22 commodities, all railroads) + cat 1 System train speed + cat 3 cars-on-line; other EP724 measures deliberately not archived in v1",
        "n_weeks": len(parsed["weeks"]),
        "n_series": len(parsed["series"]),
        "weeks": parsed["weeks"],
        "series": dict(sorted(parsed["series"].items())),
    }
    with open(OUT, "w") as fh:
        json.dump(doc, fh, separators=(",", ":"))
    print(f"wrote {os.path.relpath(OUT)} ({os.path.getsize(OUT)/1e6:.2f} MB): "
          f"{doc['n_series']} series x {doc['n_weeks']} weeks "
          f"(latest {parsed['weeks'][-1]})")


if __name__ == "__main__":
    main()
